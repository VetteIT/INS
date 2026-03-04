"""
Dataset loader pre Italian Parkinson's Voice and Speech (ItalianPVS) databazu.
ItalianPVS je taliansky dataset s recou pacientov s Parkinsonovou chorobou,
vytvoreny na Università degli Studi di Bari, Taliansko.

Zdroj: https://ieee-dataport.org/open-access/italian-parkinsons-voice-and-speech
DOI: 10.21227/aw6b-tg17
Licencia: Open Access (vyzaduje sa bezplatny IEEE ucet)

Citacia:
    1. G. Dimauro, D. Caivano, V. Bevilacqua, F. Girardi and V. Napoletano,
       "VoxTester, software for digital evaluation of speech changes in Parkinson
       disease," 2016 IEEE MeMeA, pp. 1-6. doi: 10.1109/MeMeA.2016.7533761.
    2. G. Dimauro, V. Di Nicola, V. Bevilacqua, D. Caivano and F. Girardi,
       "Assessment of Speech Intelligibility in Parkinson's Disease Using a
       Speech-To-Text System," IEEE Access, vol. 5, pp. 22199-22208, 2017.
       doi: 10.1109/ACCESS.2017.2762475.

O datasete:
    - 65 ucastnikov celkom:
      * 15 Young Healthy Controls (mlady zdravy kontroly)
      * 22 Elderly Healthy Controls (starsi zdravy kontroly)
      * 28 People with Parkinson's Disease (pacienti s PD)
    - Format: WAV audio subory + XLSX metadata
    - Velkost: ~565 MB
    - Jazyk: Taliancina
    - Recove ulohy: citanie textu, spontanna rec

Ocakavane struktury (loader podporuje viacero moznosti):

    Struktura 1 - po skupinach:
        ItalianPVS/
            Young_HC/
                *.wav
            Elderly_HC/
                *.wav
            PD/
                *.wav

    Struktura 2 - jednoducha HC/PD:
        ItalianPVS/
            HC/ (alebo healthy/ alebo controls/)
                *.wav
            PD/ (alebo parkinsons/ alebo patients/)
                *.wav

    Struktura 3 - flat s metadata:
        ItalianPVS/
            metadata.xlsx
            *.wav (vsetky subory v jednom priecinku)

Autori: Dmytro Protsun, Mykyta Olym
"""

import os
import re
import glob

from data.base_dataset import BaseAudioDataset
from config.settings import ITALIAN_PVS_DIR


class ItalianPVSDataset(BaseAudioDataset):
    """
    Dataset trieda pre Italian Parkinson's Voice and Speech databazu.
    Obsahuje nahravky od PD pacientov a zdravych kontrol z Bari, Taliansko.
    Zaznamovane v talianskom jazyku.
    """

    def __init__(self, transform=None, feature_type="spectrogram"):
        """
        Parametre:
            transform: transformacie na audio
            feature_type: typ features ("spectrogram", "mfcc", "raw")
        """
        super().__init__(
            data_dir=ITALIAN_PVS_DIR,
            domain_name="ItalianPVS",
            transform=transform,
            feature_type=feature_type
        )

    def _load_metadata(self):
        """
        Nacita zoznam audio suborov z ItalianPVS datasetu.
        Podporuje viacero moznych struktur:
        1. Tri skupiny: Young_HC/, Elderly_HC/, PD/
        2. Dve skupiny: HC/ (alebo healthy/) a PD/ (alebo parkinsons/)
        3. Flat struktura s XLSX metadata
        4. Rekurzivny scan s detekciou labelu z nazvu suboru/priecinka
        """
        if not os.path.exists(self.data_dir):
            print(f"  VAROVANIE: Priecinok {self.data_dir} neexistuje!")
            print(f"  Prosim stiahnite ItalianPVS dataset z:")
            print(f"  https://ieee-dataport.org/open-access/italian-parkinsons-voice-and-speech")
            print(f"  (Potrebujete bezplatny IEEE ucet)")
            print(f"  Rozbalte ZIP a umiestnite obsah do: {self.data_dir}")
            return

        # Najdeme skutocny root - moze byt vnoreny v podpriecinku
        root_dir = self._find_root_dir()

        # 1. Struktura s tromi skupinami (Young_HC, Elderly_HC, PD)
        if self._try_three_groups(root_dir):
            return

        # 2. Struktura s dvomi skupinami (HC/PD alebo healthy/parkinsons)
        if self._try_two_groups(root_dir):
            return

        # 3. Flat struktura - skusime XLSX metadata
        if self._try_xlsx_metadata(root_dir):
            return

        # 4. Fallback - rekurzivny scan a detekcia z nazvu
        self._load_from_filenames(root_dir)

    def _find_root_dir(self):
        """
        Najde skutocny korenovy priecinok s datami.
        Moze byt priamo data_dir alebo vnoreny podpriecinok.
        """
        # Skontrolujeme ci data_dir priamo obsahuje WAV alebo podpriecinky s WAV
        has_wav = len(glob.glob(os.path.join(self.data_dir, "*.wav"))) > 0
        has_subdirs = any(
            os.path.isdir(os.path.join(self.data_dir, d))
            for d in os.listdir(self.data_dir)
            if not d.startswith(".")
        ) if os.path.exists(self.data_dir) else False

        if has_wav or has_subdirs:
            # Ak je len jeden podpriecinok a ziadne WAV, skusime vojst hlbsie
            if not has_wav and has_subdirs:
                subdirs = [
                    d for d in os.listdir(self.data_dir)
                    if os.path.isdir(os.path.join(self.data_dir, d))
                    and not d.startswith(".")
                ]
                if len(subdirs) == 1:
                    return os.path.join(self.data_dir, subdirs[0])
            return self.data_dir

        return self.data_dir

    def _try_three_groups(self, root_dir):
        """
        Skusi nacitat data z trojskupinovej struktury.
        Hlada priecinky ako: Young_HC, Elderly_HC (alebo Old_HC), PD (alebo Parkinson)
        """
        # Hladame PD priecinok
        pd_dir = None
        hc_dirs = []

        for name in os.listdir(root_dir) if os.path.exists(root_dir) else []:
            full_path = os.path.join(root_dir, name)
            if not os.path.isdir(full_path):
                continue

            name_upper = name.upper()

            if "PD" in name_upper or "PARKINSON" in name_upper:
                pd_dir = full_path
            elif ("HC" in name_upper or "HEALTHY" in name_upper or
                  "CONTROL" in name_upper or "NORMAL" in name_upper):
                hc_dirs.append(full_path)

        # Ak mame PD a aspon 2 HC priecinky, je to trojskupinova struktura
        if pd_dir and len(hc_dirs) >= 2:
            for hc_dir in hc_dirs:
                self._load_from_dir(hc_dir, label=0, group_name=os.path.basename(hc_dir))
            self._load_from_dir(pd_dir, label=1, group_name="PD")
            return len(self.audio_paths) > 0

        return False

    def _try_two_groups(self, root_dir):
        """
        Skusi nacitat data z dvojskupinovej struktury (HC/PD).
        """
        pd_dir = None
        hc_dir = None

        for name in os.listdir(root_dir) if os.path.exists(root_dir) else []:
            full_path = os.path.join(root_dir, name)
            if not os.path.isdir(full_path):
                continue

            name_upper = name.upper()

            if "PD" in name_upper or "PARKINSON" in name_upper or "PATIENT" in name_upper:
                pd_dir = full_path
            elif ("HC" in name_upper or "HEALTHY" in name_upper or
                  "CONTROL" in name_upper or "NORMAL" in name_upper):
                hc_dir = full_path

        if pd_dir and hc_dir:
            self._load_from_dir(hc_dir, label=0, group_name="HC")
            self._load_from_dir(pd_dir, label=1, group_name="PD")
            return len(self.audio_paths) > 0

        return False

    def _try_xlsx_metadata(self, root_dir):
        """
        Skusi nacitat data pomocou XLSX metadata suboru.
        Dataset obsahuje xlsx subory s informaciami o pacientoch.
        """
        try:
            import openpyxl
        except ImportError:
            return False

        # Hladame XLSX subory
        xlsx_files = glob.glob(os.path.join(root_dir, "*.xlsx"))
        if not xlsx_files:
            xlsx_files = glob.glob(os.path.join(root_dir, "**", "*.xlsx"), recursive=True)

        if not xlsx_files:
            return False

        # Skusime kazdy XLSX subor
        for xlsx_path in xlsx_files:
            try:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        continue

                    # Hladame stlpce s nazvom suboru a labelom
                    header = [str(h).lower() if h else "" for h in rows[0]]

                    file_col = None
                    label_col = None
                    for i, h in enumerate(header):
                        if "file" in h or "wav" in h or "audio" in h or "name" in h:
                            file_col = i
                        if "label" in h or "class" in h or "group" in h or "status" in h:
                            label_col = i

                    if file_col is not None and label_col is not None:
                        for row in rows[1:]:
                            if len(row) <= max(file_col, label_col):
                                continue

                            filename = str(row[file_col]) if row[file_col] else ""
                            label_val = str(row[label_col]).upper() if row[label_col] else ""

                            if not filename:
                                continue

                            # Urcime label
                            if "PD" in label_val or "PARKINSON" in label_val or label_val == "1":
                                label = 1
                            elif "HC" in label_val or "HEALTHY" in label_val or label_val == "0":
                                label = 0
                            else:
                                continue

                            # Najdeme audio subor
                            audio_path = self._find_audio_file(root_dir, filename)
                            if audio_path:
                                self.audio_paths.append(audio_path)
                                self.labels.append(label)
                                speaker_id = self._extract_speaker_id(filename)
                                self.speaker_ids.append(speaker_id)

                wb.close()
            except Exception:
                continue

        return len(self.audio_paths) > 0

    def _find_audio_file(self, root_dir, filename):
        """Najde audio subor v priecinku alebo podpriecinkoch."""
        # Skusime priamu cestu
        if not filename.endswith(".wav"):
            filename = filename + ".wav"

        direct = os.path.join(root_dir, filename)
        if os.path.exists(direct):
            return direct

        # Hladame rekurzivne
        matches = glob.glob(os.path.join(root_dir, "**", filename), recursive=True)
        if matches:
            return matches[0]

        return None

    def _load_from_filenames(self, root_dir):
        """
        Fallback - nacita vsetky audio subory a urcuje label z nazvu suboru
        alebo nazvu nadradeneho priecinka.
        """
        all_files = []
        for ext in ["*.wav", "*.WAV"]:
            all_files.extend(
                glob.glob(os.path.join(root_dir, "**", ext), recursive=True)
            )

        for audio_file in sorted(all_files):
            # Urcime label z cesty
            parent_name = os.path.basename(os.path.dirname(audio_file)).upper()
            filename = os.path.basename(audio_file).upper()

            if ("PD" in parent_name or "PARKINSON" in parent_name or
                "PD" in filename or "PARKINSON" in filename):
                label = 1
            elif ("HC" in parent_name or "HEALTHY" in parent_name or
                  "CONTROL" in parent_name or "NORMAL" in parent_name or
                  "HC" in filename or "HEALTHY" in filename):
                label = 0
            else:
                # Ak nevieme urcit z nazvu, preskocime
                continue

            self.audio_paths.append(audio_file)
            self.labels.append(label)
            speaker_id = self._extract_speaker_id(os.path.basename(audio_file))
            self.speaker_ids.append(speaker_id)

        if len(self.audio_paths) == 0:
            print(f"  VAROVANIE: Nenasli sa ziadne audio subory v {root_dir}")
            print(f"  Skontrolujte strukturu priecinkov.")

    def _load_from_dir(self, directory, label, group_name=""):
        """
        Nacita vsetky WAV subory z daneho priecinka a jeho podpriecinkov.

        Parametre:
            directory: cesta k priecinku s WAV subormi
            label: 0 = healthy, 1 = PD
            group_name: nazov skupiny (pre speaker ID prefix)
        """
        dir_basename = os.path.basename(directory)

        for ext in ["*.wav", "*.WAV"]:
            # Hladame aj v podpriecinkoch (per-speaker priecinky)
            files = glob.glob(os.path.join(directory, "**", ext), recursive=True)
            for audio_file in sorted(files):
                self.audio_paths.append(audio_file)
                self.labels.append(label)

                # Speaker ID z nadradeneho priecinka (per-speaker priecinky)
                parent_dir = os.path.basename(os.path.dirname(audio_file))
                if parent_dir != dir_basename:
                    # Subor je v speaker podpriecinku (napr. AGNESE P/)
                    speaker_id = f"IPVS_{parent_dir}"
                else:
                    # Subor je priamo v group priecinku (flat struktura)
                    speaker_id = self._extract_speaker_id(
                        os.path.basename(audio_file), group_name
                    )
                self.speaker_ids.append(speaker_id)

    def _extract_speaker_id(self, filename, group_name=""):
        """
        Extrahuje speaker ID z nazvu suboru.
        Hlada ciselnu cast v nazve suboru.

        Priklad: PD001_read.wav -> IPVS_PD_001
                 HC_02_sentence.wav -> IPVS_HC_02
        """
        # Skusime najst ciselne ID v nazve suboru
        match = re.search(r"(\d+)", filename)
        prefix = "IPVS"

        if group_name:
            prefix = f"IPVS_{group_name}"

        if match:
            return f"{prefix}_{match.group(1)}"
        else:
            # Fallback - pouzijeme nazov suboru
            name_without_ext = os.path.splitext(filename)[0]
            return f"{prefix}_{name_without_ext}"
